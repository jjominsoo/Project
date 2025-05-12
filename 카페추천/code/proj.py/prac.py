from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import pyautogui
import openpyxl

keyword = pyautogui.prompt("검색어를 입력하십시오")
wb = openpyxl.Workbook()
ws = wb.create_sheet()
ws.append(["영업중", "가게명", "특징1", "특징1수", "특징2", "특징2수", "특징3", "특징3수", "특징4", "특징4수", "특징5", "특징5수", "특징6", "특징6수", "특징7", "특징7수", 
"특징8", "특징8수", "특징9", "특징9수", "특징10", "특징10수", "특징11", "특징11수", "특징12", "특징12수", "특징13", "특징13수", "특징14", "특징14수", "특징15", "특징15수", 
"특징16", "특징16수", "특징17", "특징17수", "특징18", "특징18수", "특징19", "특징19수", "특징20", "특징20수", "특징21", "특징21수", "특징22", "특징22수", "특징23", "특징23수" ])

browser = webdriver.Chrome("./chromedriver.exe")
browser.get("https://map.naver.com/v5/")
browser.implicitly_wait(10)
browser.maximize_window()

search = browser.find_element_by_css_selector("input.input_search")
search.click()
time.sleep(1)
search.send_keys("강남역 카페")
time.sleep(1)
search.send_keys(Keys.ENTER)
time.sleep(2)

browser.switch_to.frame("searchIframe")
browser.find_element_by_css_selector("#_pcmap_list_scroll_container").click()

def data_processing(str):
    if str.find("맛있어요") != -1:
        return 0
    elif str.find("인테리어") != -1:
        return 1
    elif str.find("사진") != -1:
        return 1    
    elif str.find("뷰") != -1:
        return 1
    elif str.find("주차") != -1:
        return 1
    elif str.find("대화") != -1:
        return 2
    elif str.find("좌석") != -1:
        return 2
    elif str.find("가성비") != -1:
        return 2
    elif str.find("집중") != -1:
        return 2
    # else:
        # str = 3

page=2
num = 0
num2 = 0
while page < 8:

    before_len = 0
    after_len = 0
    if (len(browser.find_elements_by_css_selector("li.UEzoS")) > 0):
        time.sleep(1.5)
        lis = browser.find_elements_by_css_selector("li.UEzoS")
        before_len = len(lis)

        while True:
            
            browser.switch_to.default_content()
            browser.switch_to.frame("searchIframe")
            time.sleep(1)
            browser.find_element_by_css_selector("body").send_keys(Keys.END)
            # browser.execute_script("window.scrollTo(0,document.body.scrollHeight)") 
            time.sleep(1)
            lis = browser.find_elements_by_css_selector("li.UEzoS")
            after_len = len(lis)
            print("스크롤 전 ", before_len, "/ 스크롤 후 ", after_len)
            if before_len == after_len:
                break
            before_len = after_len
        browser.implicitly_wait(0)

    else:       
        lis = browser.find_elements_by_css_selector("li.VLTHu")
        before_len = len(lis)
        while True:
            browser.find_element_by_css_selector("body").send_keys(Keys.END)
            time.sleep(1.5)
            lis = browser.find_elements_by_css_selector("li.VLTHu")
            after_len = len(lis)
            print("스크롤 전 ", before_len, "/ 스크롤 후 ", after_len)
            if before_len == after_len:
                break
            before_len = after_len
        browser.implicitly_wait(0)
    time.sleep(2)

    

    count = 0
    for li in lis:
        if len(li.find_elements_by_css_selector("svg.dPXjn")) > 0:
            count += 1
    print("광고수 = ", count)
    

    ad = 0
    for i, li in enumerate(lis, start=0):
        
        # 광고는 크롤링안함 (중복되기 때문)
        if len(li.find_elements_by_css_selector("svg.dPXjn")) > 0:
            continue
            # browser.find_element_by_class_name("N_KDL").click()


        # 현재 영업중인지? stat 0 : 영업안함 1 : 영업중
        
            
        # 세부페이지 열고
        li.find_element_by_class_name("N_KDL").click()

        time.sleep(1)
        browser.switch_to.default_content()
        browser.switch_to.frame("entryIframe")
        time.sleep(1)
        
        if (len(browser.find_elements_by_xpath('//*[@id="app-root"]/div/div/div/div[6]/div/div[2]/div/ul/li[3]/div/a/div/div/div/em')) > 0):       
            cur_stat = browser.find_element_by_xpath('//*[@id="app-root"]/div/div/div/div[6]/div/div[2]/div/ul/li[3]/div/a/div/div/div/em')
            print("cur_stat = ",cur_stat)

            if cur_stat == '영업 중':
                stat = 1
            else: 
                stat = 0

        name = browser.find_element_by_css_selector("span.Fc1rA").text
        # print("stat = ", stat, "name = ", name)

        # ws.cell(row=int(i),column=1,value=stat)
        # ws.cell(row=int(i),column=2,value=name)   
        print(name)
        
        # 메뉴탭이 4~6개까지 많다. 그때의 '리뷰'라는 탭의 위치
        # 사실 텍스트 받아서 '리뷰' 탭을 받고 싶은데 안됐다.
        tabs = browser.find_elements_by_css_selector("a.tpj9w")
        for tab in tabs:
            # print(tab.text)
            if tab.text == '리뷰':
                tab.click()
                time.sleep(1)
                break


        # if len(browser.find_elements_by_css_selector("a.tpj9w")) >  3:     
        #     if len(browser.find_elements_by_css_selector("a.tpj9w")) >  4:34
        #         if len(browser.find_elements_by_css_selector("a.tpj9w")) >  5:
        #             browser.find_element_by_css_selector("a.tpj9w:nth-child(5)").click()
        #         else:
        #             browser.find_element_by_css_selector("a.tpj9w:nth-child(4)").click()
        #     else:
        #         browser.find_element_by_css_selector("a.tpj9w:nth-child(3)").click()
        # else: 
        #     browser.find_element_by_css_selector("a.tpj9w:nth-child(2)").click()
        # time.sleep(1)
        # browser.switch_to.default_content()
                # if browser.find_element_by_css_selector("a.tpj9w:nth-child(",i+1,")") == "리뷰":
                #     browser.find_element_by_css_selector("a.tpj9w:nth-child(",i+1,")").click()

        # for i in range(len(browser.find_elements_by_css_selector("a.q1Lf9"))):

        # '리뷰'탭 클릭하고 나오는 더보기란 클릭
        if(len(browser.find_elements_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]')))>0:
            browser.find_element_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]').click()
        time.sleep(1)

        # 만약 더보기가 더 있다면 계속 클릭
        while True:
            if len(browser.find_elements_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[2]')) > 0:
                # print(len(browser.find_elements_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]')))
                browser.find_element_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]').click()
            else:
                break
        # browser.find_element_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]').click()
    
        
        features = browser.find_elements_by_css_selector("li.nbD78")
        # print("features = ",len(features))
        # if len(features) < 10:
        #     browser.switch_to.default_content()
        #     browser.switch_to.frame("searchIframe")
        #     # ws.append(" ")
        #     continue

        # if len(browser.find_elements_by_xpath('//*[@id="app-root"]/div/div/div/div[7]/div[2]/div[1]/div/div/div/div')) >0:
        ## 리뷰가 없다면
        if browser.find_elements_by_class_name('Zi4oC'):
            browser.switch_to.default_content()
            browser.switch_to.frame("searchIframe")
            ad += 1
            # ws.append(" ")
            continue

        for j, feature in enumerate(features,start=2):

            # print(feature)
            ##################################################################### 리뷰가 없는경우 처리
            f_id = feature.find_element_by_css_selector("span.nWiXa").text
            f_num = feature.find_element_by_css_selector("span.TwM9q").text
            f_id = f_id.replace("\"", "")

            new_f_id = data_processing(f_id)
            # print(f_id)
            f_num = f_num.replace("이 키워드를 선택한 인원\n","")
            f_num = int(f_num)

            ##################################################################### 여기 페이지수 다시
            if int(i)+int(num)+2-ad < 0:
                ws.cell(row=int(i)+int(num)+2,column=1,value=stat)
                ws.cell(row=int(i)+int(num)+2,column=2,value=name)
                ws.cell(row=int(i)+int(num)+2,column=int(j)*2-1,value=f_id)
                # ws.cell(row=int(i)+int(num)+2,column=int(j)*2-1,value=new_f_id)
                ws.cell(row=int(i)+int(num)+2,column=int(j)*2,value=f_num)
            ws.cell(row=int(i)+int(num)+2-ad,column=1,value=stat)
            ws.cell(row=int(i)+int(num)+2-ad,column=2,value=name)
            ws.cell(row=int(i)+int(num)+2-ad,column=int(j)*2-1,value=f_id)
            # ws.cell(row=int(i)+int(num)+2,column=int(j)*2-1,value=new_f_id)
            ws.cell(row=int(i)+int(num)+2-ad,column=int(j)*2,value=f_num)
            # print("f_id = ", f_id, "/ f_num = ", f_num)
        time.sleep(1)
        browser.switch_to.default_content()
        browser.switch_to.frame("searchIframe")
            # browser.find_element_by_link_text(str(page)).click

    pages = browser.find_elements_by_css_selector("a.mBN2s")


    for i in  pages:
        if int(i.text) == page:
            print("page = ",page)
            i.click()
            time.sleep(2)
            flag = 1
            
            break
        flag = 0
    print("flag = ",flag)
    if flag == 0:
        print("no pages")
        break
    time.sleep(2)
    page = page+1
    num += after_len - count


wb.save(f"proj.py/{keyword}.xlsx")
    
# names = browser.find_elements_by_class_name("N_KDL") 
# for name in names:
#     print(name.text)