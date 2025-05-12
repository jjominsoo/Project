from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import pyautogui
import openpyxl

keyword = pyautogui.prompt("검색어를 입력하십시오")
wb = openpyxl.Workbook()
ws = wb.create_sheet()
ws.append(["영업중", "가게명", "특징1", "특징1수", "특징2", "특징2수", "특징3", "특징3수", "특징4", "특징4수", "특징5", "특징5수", "특징6", "특징6수", "특징7", "특징7수", 
"특징8", "특징8수", "특징9", "특징9수", "특징10", "특징10수", "특징11", "특징11수", "특징12", "특징12수", "특징13", "특징13수", "특징14", "특징14수", "특징15", "특징15수", 
"특징16", "특징16수", "특징17", "특징17수", "특징18", "특징18수", "특징19", "특징19수", "특징20", "특징20수", "특징21", "특징21수", "특징22", "특징22수", "특징23", "특징23수" ])

browser = webdriver.Chrome("./chromedriver.exe")
browser.get("https://map.naver.com/v5/")
browser.implicitly_wait(10)
browser.maximize_window()

search = browser.find_element_by_css_selector("input.input_search")
search.click()
time.sleep(1)
search.send_keys("역삼역 카페")
time.sleep(1)
search.send_keys(Keys.ENTER)
time.sleep(2)

browser.switch_to.frame("searchIframe")
browser.find_element_by_css_selector("#_pcmap_list_scroll_container").click()


lis = browser.find_elements_by_css_selector("li.UEzoS")
before_len = len(lis)

while True:
    browser.find_element_by_css_selector("body").send_keys(Keys.END)
    time.sleep(1.5)
    lis = browser.find_elements_by_css_selector("li.UEzoS")
    after_len = len(lis)
    print("스크롤 전 ", before_len, "/ 스크롤 후 ", after_len)
    if before_len == after_len:
        break
    before_len = after_len

browser.implicitly_wait(0)

def data_processing(str):
    if str.find("맛있어요") != -1:
        return 0
    elif str.find("인테리어") != -1:
        return 1
    elif str.find("사진") != -1:
        return 1    
    elif str.find("뷰") != -1:
        return 1
    elif str.find("주차") != -1:
        return 1
    elif str.find("대화") != -1:
        return 2
    elif str.find("좌석") != -1:
        return 2
    elif str.find("가성비") != -1:
        return 2
    elif str.find("집중") != -1:
        return 2
    # else:
        # str = 3



count = 0
for i, li in enumerate(lis, start=0):

    if len(li.find_elements_by_css_selector("svg.dPXjn")) > 0:
        count = 1
        # browser.find_element_by_class_name("N_KDL").click()
    if count == 1:
        count = 0
        continue
        # 현재 영업중인지? stat 0 : 영업안함 1 : 영업중
    cur_stat = li.find_element_by_css_selector("span.h69bs.KvAhC").text
    if cur_stat == '영업 중':
        stat = 1
    else: 
        stat = 0    
        
    li.find_element_by_class_name("N_KDL").click()

    time.sleep(1)
    browser.switch_to.default_content()
    browser.switch_to.frame("entryIframe")
    name = browser.find_element_by_css_selector("span.Fc1rA").text
    # print("stat = ", stat, "name = ", name)

    # ws.cell(row=int(i),column=1,value=stat)
    # ws.cell(row=int(i),column=2,value=name)

    ws.append([stat, name])
   
    

    if len(browser.find_elements_by_css_selector("a.tpj9w")) >  4:     
        if len(browser.find_elements_by_css_selector("a.tpj9w")) >  5:
            browser.find_element_by_css_selector("a.tpj9w:nth-child(5)").click()
        else:
            browser.find_element_by_css_selector("a.tpj9w:nth-child(4)").click()
    else:
        browser.find_element_by_css_selector("a.tpj9w:nth-child(3)").click()
    temp = 3
    time.sleep(1)
    # browser.switch_to.default_content()
            # if browser.find_element_by_css_selector("a.tpj9w:nth-child(",i+1,")") == "리뷰":
            #     browser.find_element_by_css_selector("a.tpj9w:nth-child(",i+1,")").click()

    # for i in range(len(browser.find_elements_by_css_selector("a.q1Lf9"))):
    if(len(browser.find_elements_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]')))>0:
        browser.find_element_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]').click()
    time.sleep(1)
    while True:
        if len(browser.find_elements_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[2]')) > 0:
            # print(len(browser.find_elements_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]')))
            browser.find_element_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]').click()
        else:
            break
    # browser.find_element_by_xpath('/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div/div/div[2]/a[1]').click()
    ## 더보기 클릭이 안된다
        
    features = browser.find_elements_by_css_selector("li.nbD78")
    for j, feature in enumerate(features,start=2):
            # print(feature.text)
        f_id = feature.find_element_by_css_selector("span.nWiXa").text
        f_num = feature.find_element_by_css_selector("span.TwM9q").text
        f_id = f_id.replace("\"", "")
        new_f_id = data_processing(f_id)
        print(new_f_id)
        f_num = f_num.replace("이 키워드를 선택한 인원\n","")
        f_num = int(f_num)
        ws.cell(row=int(i),column=int(j)*2-1,value=f_id)
        ws.cell(row=int(i),column=int(j)*2-1,value=new_f_id)
        ws.cell(row=int(i),column=int(j)*2,value=f_num)
        # print("f_id = ", f_id, "/ f_num = ", f_num)
        
    browser.switch_to.default_content()
    browser.switch_to.frame("searchIframe")

wb.save(f"proj.py/{keyword}.xlsx")
    
# names = browser.find_elements_by_class_name("N_KDL") 
# for name in names:
#     print(name.text)